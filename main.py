from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
import shutil
import os
import openpyxl
import unicodedata
import re
from datetime import datetime

app = FastAPI(title="Agente Fazenda")

PLANILHA_PATH = "/data/Planilha_fazenda.xlsm"
BACKUP_DIR = "/data/backups"
ABAS_IGNORADAS = {"PAGINA INICIAL", "PÁGINA INICIAL", "GERAL", "BASE", "LOG"}


class RunRequest(BaseModel):
    text: str


@app.get("/")
def home():
    return {"status": "gado-agent rodando"}


def garantir_pastas():
    os.makedirs("/data", exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)


def normalizar(texto: str) -> str:
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto


def carregar_wb():
    if not os.path.exists(PLANILHA_PATH):
        raise FileNotFoundError(f"Planilha não encontrada em {PLANILHA_PATH}")
    return openpyxl.load_workbook(PLANILHA_PATH, keep_vba=True, data_only=False)


def listar_fazendas():
    wb = carregar_wb()
    ignoradas_norm = {normalizar(x) for x in ABAS_IGNORADAS}
    resultado = []
    for aba in wb.sheetnames:
        if normalizar(aba) not in ignoradas_norm:
            resultado.append(aba)
    return resultado


def encontrar_linha_header(ws, max_linhas=80):
    for r in range(1, max_linhas + 1):
        valor = ws.cell(r, 2).value
        if isinstance(valor, str):
            t = normalizar(valor).replace(" ", "")
            if "tipo" in t and "piquete" in t:
                return r
    raise ValueError(f"Header 'tipo \\\\ piquete' não encontrado na aba {ws.title}")


def construir_indice_fazenda(fazenda: str):
    wb = carregar_wb()

    if fazenda not in wb.sheetnames:
        raise ValueError(f"Fazenda '{fazenda}' não encontrada.")

    ws = wb[fazenda]
    header_row = encontrar_linha_header(ws)

    piquetes = []
    col = 3
    while True:
        valor = ws.cell(header_row, col).value
        if valor is None:
            break

        nome = str(valor).strip()
        if normalizar(nome) == "total":
            break

        piquetes.append((nome, col))
        col += 1

    categorias = []
    row = header_row + 1
    while True:
        valor = ws.cell(row, 2).value
        if valor is None or str(valor).strip() == "":
            break

        categorias.append((str(valor).strip(), row))
        row += 1

    return {
        "ws": ws,
        "header_row": header_row,
        "piquetes": piquetes,
        "categorias": categorias,
    }


def escolher_fazenda(nome_digitado: str):
    nome_digitado_n = normalizar(nome_digitado)
    fazendas = listar_fazendas()

    for f in fazendas:
        if normalizar(f) == nome_digitado_n:
            return f

    for f in fazendas:
        if nome_digitado_n in normalizar(f):
            return f

    return None


def escolher_categoria(fazenda: str, nome_digitado: str):
    idx = construir_indice_fazenda(fazenda)
    nome_n = normalizar(nome_digitado)

    for cat, row in idx["categorias"]:
        if normalizar(cat) == nome_n:
            return cat, row

    alternativas = [
        nome_n.rstrip("s"),
        nome_n + "s",
    ]
    for alt in alternativas:
        for cat, row in idx["categorias"]:
            if normalizar(cat) == alt:
                return cat, row

    return None, None


def escolher_piquete(fazenda: str, nome_digitado: str):
    idx = construir_indice_fazenda(fazenda)
    nome_n = normalizar(nome_digitado)

    for piq, col in idx["piquetes"]:
        if normalizar(piq) == nome_n:
            return piq, col

    for piq, col in idx["piquetes"]:
        if nome_n in normalizar(piq):
            return piq, col

    return None, None


def listar_piquetes(fazenda: str):
    idx = construir_indice_fazenda(fazenda)
    return [p for p, _ in idx["piquetes"]]


def listar_categorias(fazenda: str):
    idx = construir_indice_fazenda(fazenda)
    return [c for c, _ in idx["categorias"]]


def total_categoria_na_fazenda(fazenda: str, categoria: str):
    idx = construir_indice_fazenda(fazenda)
    cat_real, row = escolher_categoria(fazenda, categoria)
    if not row:
        raise ValueError(f"Categoria '{categoria}' não encontrada em {fazenda}.")

    total = 0
    detalhes = []
    for piq, col in idx["piquetes"]:
        valor = idx["ws"].cell(row, col).value
        qtd = int(valor or 0)
        total += qtd
        detalhes.append((piq, qtd))

    return cat_real, total, detalhes


def quantidade_categoria_no_piquete(fazenda: str, categoria: str, piquete: str):
    idx = construir_indice_fazenda(fazenda)
    cat_real, row = escolher_categoria(fazenda, categoria)
    piq_real, col = escolher_piquete(fazenda, piquete)

    if not row:
        raise ValueError(f"Categoria '{categoria}' não encontrada em {fazenda}.")
    if not col:
        raise ValueError(f"Piquete '{piquete}' não encontrado em {fazenda}.")

    valor = idx["ws"].cell(row, col).value
    return cat_real, piq_real, int(valor or 0)


def criar_backup():
    garantir_pastas()
    if not os.path.exists(PLANILHA_PATH):
        return None

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_path = os.path.join(BACKUP_DIR, f"Planilha_fazenda_backup_{timestamp}.xlsm")
    shutil.copy2(PLANILHA_PATH, backup_path)
    return backup_path


@app.post("/run")
def run(req: RunRequest):
    texto = req.text.strip()
    texto_n = normalizar(texto)

    try:
        if texto_n == "fazendas":
            fazendas = listar_fazendas()
            return {"reply": "Fazendas: " + ", ".join(fazendas)}

        if texto_n.startswith("piquetes "):
            nome_fazenda = texto[9:].strip()
            fazenda_real = escolher_fazenda(nome_fazenda)

            if not fazenda_real:
                return {"reply": f"Não encontrei a fazenda '{nome_fazenda}'."}

            piquetes = listar_piquetes(fazenda_real)
            return {"reply": f"Piquetes de {fazenda_real}: " + ", ".join(piquetes)}

        if texto_n.startswith("categorias "):
            nome_fazenda = texto[11:].strip()
            fazenda_real = escolher_fazenda(nome_fazenda)

            if not fazenda_real:
                return {"reply": f"Não encontrei a fazenda '{nome_fazenda}'."}

            cats = listar_categorias(fazenda_real)
            return {"reply": f"Categorias de {fazenda_real}: " + ", ".join(cats)}

        m = re.match(r"quantos?\s+(.+?)\s+tem\s+no?\s+(.+)$", texto_n)
        if m:
            categoria_txt = m.group(1).strip()
            fazenda_txt = m.group(2).strip()

            fazenda_real = escolher_fazenda(fazenda_txt)
            if not fazenda_real:
                return {"reply": f"Não encontrei a fazenda '{fazenda_txt}'."}

            cat_real, total, detalhes = total_categoria_na_fazenda(fazenda_real, categoria_txt)
            detalhes_txt = ", ".join([f"{piq}: {qtd}" for piq, qtd in detalhes])
            return {
                "reply": f"Em {fazenda_real}, há {total} de {cat_real}. Distribuição: {detalhes_txt}."
            }

        m = re.match(r"quantos?\s+(.+?)\s+no\s+(.+?)\s+de\s+(.+)$", texto_n)
        if m:
            categoria_txt = m.group(1).strip()
            piquete_txt = m.group(2).strip()
            fazenda_txt = m.group(3).strip()

            fazenda_real = escolher_fazenda(fazenda_txt)
            if not fazenda_real:
                return {"reply": f"Não encontrei a fazenda '{fazenda_txt}'."}

            cat_real, piq_real, qtd = quantidade_categoria_no_piquete(
                fazenda_real, categoria_txt, piquete_txt
            )
            return {
                "reply": f"No piquete {piq_real} da fazenda {fazenda_real}, há {qtd} de {cat_real}."
            }

        return {
            "reply": (
                "Comandos disponíveis: 'fazendas', 'piquetes limão', 'categorias limão', "
                "'quantos bois tem no limão', 'quantos bois no coxo azul 2 de limão', "
                "'me manda a planilha'."
            )
        }

    except Exception as e:
        return {"reply": f"Erro: {str(e)}"}


@app.get("/download-planilha")
def download_planilha():
    if not os.path.exists(PLANILHA_PATH):
        raise HTTPException(status_code=404, detail="Planilha não encontrada.")

    nome_download = f"Planilha_fazenda_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsm"
    return FileResponse(
        path=PLANILHA_PATH,
        filename=nome_download,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
    )


@app.post("/upload")
async def upload_planilha(file: UploadFile = File(...)):
    garantir_pastas()

    if not file.filename:
        raise HTTPException(status_code=400, detail="Nenhum arquivo enviado.")

    if not file.filename.lower().endswith(".xlsm"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .xlsm")

    backup_path = criar_backup()

    with open(PLANILHA_PATH, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return {
        "status": "planilha enviada com sucesso",
        "filename": file.filename,
        "path": PLANILHA_PATH,
        "backup": backup_path
    }
